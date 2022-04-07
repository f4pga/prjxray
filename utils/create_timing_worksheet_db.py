#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Copyright (C) 2017-2020  The Project X-Ray Authors.
#
# Use of this source code is governed by a ISC-style
# license that can be found in the LICENSE file or at
# https://opensource.org/licenses/ISC
#
# SPDX-License-Identifier: ISC
""" This takes a JSON file generated with write_timing_info.tcl and generates
a spreadsheet with the prjxray timing model and compares it with the
interconnect timing output from Vivado.

"""
import argparse
import json
from openpyxl import Workbook, utils
from prjxray.tile import OutPinTiming, InPinTiming
from prjxray.timing import Outpin, Inpin, Wire, Buffer, \
        PassTransistor, IntristicDelay, RcElement, PvtCorner
from prjxray.math_models import ExcelMathModel
from prjxray.db import Database
from prjxray.util import OpenSafeFile
from prjxray import util

FAST = PvtCorner.FAST
SLOW = PvtCorner.SLOW


class TimingLookup(object):
    def __init__(self, db, nodes):
        self.db = db
        self.grid = db.grid()
        self.nodes = nodes

    def try_find_site_pin(self, site_pin_node, node_idx):
        site_pin_wire = self.nodes[site_pin_node]['wires'][node_idx]['name']
        tile, wire_in_tile = site_pin_wire.split('/')

        gridinfo = self.grid.gridinfo_at_tilename(tile)

        tile_type = self.db.get_tile_type(gridinfo.tile_type)

        for site in tile_type.get_sites():
            for site_pin in site.site_pins:
                if site_pin.wire == wire_in_tile:
                    return site_pin

        return None

    def find_site_pin(self, site_pin_node, node_idx):
        site_pin = self.try_find_site_pin(site_pin_node, node_idx)

        assert site_pin is not None, site_pin_node
        return site_pin

    def find_pip(self, pip_name):
        tile, pip = pip_name.split('/')

        gridinfo = self.grid.gridinfo_at_tilename(tile)

        tile_type = self.db.get_tile_type(gridinfo.tile_type)

        return tile_type.get_pip_by_name(pip)

    def find_wire(self, wire_name):
        tile, wire_in_tile = wire_name.split('/')

        gridinfo = self.grid.gridinfo_at_tilename(tile)

        tile_type = self.db.get_tile_type(gridinfo.tile_type)

        return tile_type.wires[wire_in_tile]


def delays_to_cells(ws, row, delays, cells):
    cells['FAST_MAX'] = 'E{}'.format(row)
    cells['FAST_MIN'] = 'F{}'.format(row)
    cells['SLOW_MAX'] = 'G{}'.format(row)
    cells['SLOW_MIN'] = 'H{}'.format(row)

    if delays is not None:
        ws[cells['FAST_MAX']] = delays[FAST].max
        ws[cells['FAST_MIN']] = delays[FAST].min
        ws[cells['SLOW_MAX']] = delays[SLOW].max
        ws[cells['SLOW_MIN']] = delays[SLOW].min
    else:
        ws[cells['FAST_MAX']] = 0
        ws[cells['FAST_MIN']] = 0
        ws[cells['SLOW_MAX']] = 0
        ws[cells['SLOW_MIN']] = 0


def cells_to_delays(cells):
    return {
        FAST: IntristicDelay(min=cells['FAST_MIN'], max=cells['FAST_MAX']),
        SLOW: IntristicDelay(min=cells['SLOW_MIN'], max=cells['SLOW_MAX']),
    }


class Net(object):
    def __init__(self, net):
        self.net = net
        self.ipin_nodes = {}
        self.row = None
        self.math = ExcelMathModel()
        self.models = {}

        for ipin in net['ipins']:
            for ipin_node in ipin['node'].strip().split(' '):
                self.ipin_nodes[ipin_node] = ipin

        # Map of wire name to parent node
        self.wire_to_node = {}

        # Map of node name to node
        self.node_name_to_node = {}

        for node in net['nodes']:
            self.node_name_to_node[node['name']] = node
            for wire in node['wires']:
                self.wire_to_node[wire['name']] = node

        # Map of (src node, dst wire).
        self.pips = {}
        for pip in net['pips']:
            src_node = self.wire_to_node[pip['src_wire']]['name']
            dst_wire = pip['dst_wire'].split('/')[1]
            self.pips[(src_node, dst_wire)] = pip

            if not int(pip['is_directional']):
                dst_node = self.wire_to_node[pip['dst_wire']]['name']
                src_wire = pip['src_wire'].split('/')[1]
                self.pips[(dst_node, src_wire)] = pip

    def extend_rc_tree(self, ws, current_rc_root, timing_lookup, node):
        rc_elements = []

        # LV nodes have a workaround applied because of a work around in the
        # pip timing data.
        is_lv_node = any(
            wire['name'].split('/')[1].startswith('LV')
            for wire in node['wires'])
        for idx, wire in enumerate(node['wires']):
            wire_timing = timing_lookup.find_wire(wire['name'])
            ws['A{}'.format(self.row)] = wire['name']
            ws['B{}'.format(self.row)] = 'Part of wire'

            if wire_timing is not None:
                cells = {}
                cells['R'] = 'C{}'.format(self.row)
                cells['C'] = 'D{}'.format(self.row)
                if not is_lv_node:
                    ws[cells['R']] = wire_timing.resistance
                    ws[cells['C']] = wire_timing.capacitance
                else:
                    # Only use first 2 wire RC's, ignore the rest.  It appears
                    # that some of the RC constant was lumped into the switch
                    # timing, so don't double count.
                    if idx < 2:
                        ws[cells['R']] = wire_timing.resistance
                        ws[cells['C']] = wire_timing.capacitance
                    else:
                        ws[cells['R']] = 0
                        ws[cells['C']] = 0

                rc_elements.append(
                    RcElement(
                        resistance=cells['R'],
                        capacitance=cells['C'],
                    ))

            self.row += 1

        wire_rc_node = Wire(rc_elements=rc_elements, math=self.math)
        self.models[self.row - 1] = wire_rc_node

        current_rc_root.set_sink_wire(wire_rc_node)

        return wire_rc_node

    def descend_route(
            self,
            ws,
            timing_lookup,
            current_node,
            route,
            route_idx,
            current_rc_root,
            was_opin=False):
        """ Traverse the next pip, or recurse deeper. """

        # descend_route should've consumed this token
        assert route[route_idx] != '}'

        while route[route_idx] == '{':
            # Go deeper
            route_idx = self.descend_route(
                ws,
                timing_lookup,
                current_node,
                route,
                route_idx=route_idx + 1,
                current_rc_root=current_rc_root,
                was_opin=was_opin)

        next_edge = (current_node, route[route_idx])
        route_idx += 1
        assert next_edge in self.pips, (next_edge, self.pips.keys())

        pip = self.pips[next_edge]
        is_backward = self.wire_to_node[
            pip['dst_wire']]['name'] == current_node
        if not is_backward:
            assert self.wire_to_node[
                pip['src_wire']]['name'] == current_node, (current_node, pip)

        pip_info = timing_lookup.find_pip(pip['name'])
        if not is_backward:
            pip_timing = pip_info.timing
            current_node = self.wire_to_node[pip['dst_wire']]['name']
        else:
            pip_timing = pip_info.backward_timing
            current_node = self.wire_to_node[pip['src_wire']]['name']

        ws['A{}'.format(self.row)] = pip['name']

        cells = {}
        cells['R'] = 'C{}'.format(self.row)
        cells['C'] = 'D{}'.format(self.row)
        delays_to_cells(
            ws, row=self.row, delays=pip_timing.delays, cells=cells)
        delays = cells_to_delays(cells)

        if pip_info.is_pass_transistor:
            ws['B{}'.format(self.row)] = 'PassTransistor'
            ws[cells['R']] = pip_timing.drive_resistance

            pip_model = PassTransistor(
                drive_resistance=cells['R'],
                delays=delays,
            )
        else:
            ws['B{}'.format(self.row)] = 'Buffer'
            if pip_timing.drive_resistance is not None:
                ws[cells['R']] = pip_timing.drive_resistance
                if pip_timing.drive_resistance == 0 and was_opin:
                    new_site_pin = timing_lookup.try_find_site_pin(
                        current_node, node_idx=0)
                    if new_site_pin is not None:
                        ws[cells['R']] = new_site_pin.timing.drive_resistance
            else:
                ws[cells['R']] = 0

            if pip_timing.internal_capacitance is not None:
                ws[cells['C']] = pip_timing.internal_capacitance
            else:
                ws[cells['C']] = 0

            pip_model = Buffer(
                drive_resistance=cells['R'],
                internal_capacitance=cells['C'],
                delays=cells_to_delays(cells),
            )

        self.models[self.row] = pip_model

        current_rc_root.add_child(pip_model)
        self.row += 1

        if current_node in self.ipin_nodes:
            assert route[route_idx] in ['}', 'IOB_O_OUT0', 'IOB_T_OUT0'], (
                route_idx,
                route[route_idx],
            )
            route_idx += 1

        node = self.node_name_to_node[current_node]

        current_rc_root = self.extend_rc_tree(
            ws, pip_model, timing_lookup, node)

        if current_node in self.ipin_nodes:
            ipin = self.ipin_nodes[current_node]

            cells = {}
            name = '{} to {}'.format(self.net['opin']['name'], ipin['name'])
            ws['A{}'.format(self.row)] = ipin['name']
            ws['B{}'.format(self.row)] = 'Inpin'

            site_pin = timing_lookup.find_site_pin(current_node, node_idx=-1)
            assert isinstance(site_pin.timing, InPinTiming)

            cells = {}
            cells['C'] = 'D{}'.format(self.row)
            delays_to_cells(
                ws, row=self.row, delays=site_pin.timing.delays, cells=cells)
            delays = cells_to_delays(cells)

            ws[cells['C']] = site_pin.timing.capacitance
            ipin_model = Inpin(
                capacitance=cells['C'], delays=delays, name=name)
            self.models[self.row] = ipin_model
            current_rc_root.add_child(ipin_model)
            self.row += 1

            #Sum delays only (sum*1000)
            #Sum delays + capacitive delay
            #
            #Total delay (from Vivado)
            ws['A{}'.format(self.row)] = '{}: {} sum delays'.format(
                self.net['net'], name)
            self.row += 1

            ws['A{}'.format(self.row)] = '{}: {}sum delays + cap delay'.format(
                self.net['net'], name)
            self.row += 2

            ws['A{}'.format(
                self.row)] = '{}: {}Total delay (from Vivado)'.format(
                    self.net['net'], name)

            ws['E{}'.format(self.row)] = ipin['ic_delays']['FAST_MAX']
            ws['F{}'.format(self.row)] = ipin['ic_delays']['FAST_MIN']
            ws['G{}'.format(self.row)] = ipin['ic_delays']['SLOW_MAX']
            ws['H{}'.format(self.row)] = ipin['ic_delays']['SLOW_MIN']

            self.row += 2

            return route_idx
        else:
            return self.descend_route(
                ws,
                timing_lookup,
                current_node,
                route,
                route_idx=route_idx,
                current_rc_root=current_rc_root)

    def walk_route(self, ws, timing_lookup):
        """ Walk route, creating rows in table.

        First row will always be the OPIN, followed by the node/wire connected
        to the OPIN.  After a node/wire is always 1 or more pips. After a pip
        is always a node/wire.  A terminal node/wire will then reach an IPIN.
        """

        self.row = 2
        ws['A{}'.format(self.row)] = self.net['opin']['wire']

        site_pin = timing_lookup.find_site_pin(
            self.net['opin']['node'], node_idx=0)
        assert isinstance(site_pin.timing, OutPinTiming)

        ws['B{}'.format(self.row)] = 'Outpin'
        ws['C{}'.format(self.row)] = site_pin.timing.drive_resistance

        cells = {}
        cells['R'] = 'C{}'.format(self.row)

        delays_to_cells(
            ws, row=self.row, delays=site_pin.timing.delays, cells=cells)

        model_root = Outpin(
            resistance=cells['R'], delays=cells_to_delays(cells))
        self.models[self.row] = model_root
        self.row += 1

        node = self.net['opin']['node']

        tile, first_wire = self.net['opin']['node'].split('/')

        route = [r for r in self.net['route'].strip().split(' ') if r != '']
        assert route[0] == '{'
        assert route[1] == first_wire

        node = self.node_name_to_node[node]
        current_rc_root = self.extend_rc_tree(
            ws, model_root, timing_lookup, node)

        self.descend_route(
            ws,
            timing_lookup,
            node['name'],
            route,
            route_idx=2,
            current_rc_root=current_rc_root,
            was_opin=True)

        model_root.propigate_delays(self.math)

        model_rows = {}

        for row, model in self.models.items():
            model_rows[id(model)] = row

        for row, model in self.models.items():
            rc_delay = model.get_rc_delay()
            if rc_delay is not None:
                ws['J{}'.format(row)] = self.math.eval(rc_delay)

            downstream_cap = model.get_downstream_cap()
            if downstream_cap is not None:
                ws['I{}'.format(row)] = self.math.eval(downstream_cap)

            if isinstance(model, Inpin):
                ipin_results = {
                    'Name': model.name,
                    'truth': {},
                    'computed': {},
                }

                ipin_delays = {}

                DELAY_COLS = (
                    ('E', 'FAST_MAX'),
                    ('F', 'FAST_MIN'),
                    ('G', 'SLOW_MAX'),
                    ('H', 'SLOW_MIN'),
                )

                for col, value in DELAY_COLS:
                    ipin_delays[value] = []
                    ipin_results['computed'][
                        value] = '{title}!{col}{row}'.format(
                            title=utils.quote_sheetname(ws.title),
                            col=col,
                            row=row + 2)
                    ipin_results['truth'][value] = '{title}!{col}{row}'.format(
                        title=utils.quote_sheetname(ws.title),
                        col=col,
                        row=row + 4)

                rc_delays = []

                for model in model.get_delays():
                    delays = model.get_intrinsic_delays()
                    if delays is not None:
                        ipin_delays['FAST_MAX'].append(delays[FAST].max)
                        ipin_delays['FAST_MIN'].append(delays[FAST].min)
                        ipin_delays['SLOW_MAX'].append(delays[SLOW].max)
                        ipin_delays['SLOW_MIN'].append(delays[SLOW].min)

                    if id(model) in model_rows:
                        rc_delays.append('J{}'.format(model_rows[id(model)]))

                ws['J{}'.format(row + 1)] = self.math.eval(
                    self.math.sum(rc_delays))

                for col, value in DELAY_COLS:
                    ws['{}{}'.format(col, row + 1)] = self.math.eval(
                        self.math.sum(ipin_delays[value]))
                    ws['{}{}'.format(
                        col, row + 2)] = '=1000*({col}{row} + J{row})'.format(
                            col=col, row=row + 1)

                yield ipin_results


def add_net(wb, net, timing_lookup):
    replace_underscore = str.maketrans('[]\\:/', '_____')
    ws = wb.create_sheet(
        title="Net {}".format(net['net'].translate(replace_underscore)))

    # Header
    ws['A1'] = 'Name'
    ws['B1'] = 'Type'
    ws['C1'] = 'RES'
    ws['D1'] = 'CAP'
    ws['E1'] = 'FAST_MAX'
    ws['F1'] = 'FAST_MIN'
    ws['G1'] = 'SLOW_MAX'
    ws['H1'] = 'SLOW_MIN'
    ws['I1'] = 'Downstream C'
    ws['J1'] = 'Delay from cap'

    net_obj = Net(net)
    yield from net_obj.walk_route(ws, timing_lookup)


def build_wire_filter(wire_filter):
    wires_to_include = set()

    with OpenSafeFile(wire_filter) as f:
        for l in f:
            wire = l.strip()
            if not wire:
                continue
            wires_to_include.add(wire)

    def filter_net(net):
        wires_in_net = set()

        for node in net['nodes']:
            for wire in node['wires']:
                wires_in_net.add(wire['name'])

        return len(wires_in_net & wires_to_include) > 0

    return filter_net


def main():
    parser = argparse.ArgumentParser(
        description="Create timing worksheet for 7-series timing analysis.")

    util.db_root_arg(parser)
    util.part_arg(parser)
    parser.add_argument('--timing_json', required=True)
    parser.add_argument('--output_xlsx', required=True)
    parser.add_argument(
        '--wire_filter',
        help='List of wires that must be present in a net to be output')

    args = parser.parse_args()

    with OpenSafeFile(args.timing_json) as f:
        timing = json.load(f)

    db = Database(args.db_root, args.part)

    nodes = {}
    for net in timing:
        for node in net['nodes']:
            nodes[node['name']] = node

    timing_lookup = TimingLookup(db, nodes)

    wb = Workbook()
    summary_ws = wb[wb.sheetnames[0]]
    summary_ws.title = 'Summary'

    summary_ws['A1'] = 'Name'

    cols = ['FAST_MAX', 'FAST_MIN', 'SLOW_MAX', 'SLOW_MIN']
    cur_col = 'B'
    for col in cols:
        summary_ws['{}1'.format(cur_col)] = col
        cur_col = chr(ord(cur_col) + 1)
        summary_ws['{}1'.format(cur_col)] = 'Computed ' + col
        cur_col = chr(ord(cur_col) + 3)

    if args.wire_filter:
        wire_filter = build_wire_filter(args.wire_filter)
    else:
        wire_filter = lambda x: True

    summary_row = 2

    timing = [net for net in timing if wire_filter(net)]
    for idx, net in enumerate(timing):
        if '<' in net['route']:
            print(
                "WARNING: Skipping net {} because it has complicated route description."
                .format(net['net']))
            continue

        print('Process net {} ({} / {})'.format(net['net'], idx, len(timing)))
        for summary_cells in add_net(wb, net, timing_lookup):
            summary_ws['A{}'.format(summary_row)] = summary_cells['Name']

            cur_col = 'B'
            for col in cols:
                truth_col = chr(ord(cur_col) + 0)
                computed_col = chr(ord(cur_col) + 1)
                error_col = chr(ord(cur_col) + 2)
                error_per_col = chr(ord(cur_col) + 3)
                summary_ws['{}{}'.format(
                    truth_col,
                    summary_row)] = '=' + summary_cells['truth'][col]
                summary_ws['{}{}'.format(
                    computed_col,
                    summary_row)] = '=' + summary_cells['computed'][col]
                summary_ws['{}{}'.format(
                    error_col,
                    summary_row)] = '={truth}{row}-{comp}{row}'.format(
                        truth=truth_col, comp=computed_col, row=summary_row)
                summary_ws['{}{}'.format(
                    error_per_col,
                    summary_row)] = '={error}{row}/{truth}{row}'.format(
                        error=error_col, truth=truth_col, row=summary_row)

                cur_col = chr(ord(cur_col) + 4)

            summary_row += 1

    wb.save(filename=args.output_xlsx)


if __name__ == "__main__":
    main()
